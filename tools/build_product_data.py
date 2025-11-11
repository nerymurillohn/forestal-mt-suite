#!/usr/bin/env python3
"""Builds Forestal MT product data artifacts from the Excel masters."""
from __future__ import annotations

import csv
import json
import re
from collections import OrderedDict
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
PRODUCTS_DIR = ROOT / "products-data"
CATALOG_XLSX = PRODUCTS_DIR / "forestal-mt-products-catalogue-46-skus.xlsx"
RETAIL_XLSX = PRODUCTS_DIR / "forestal-mt-products-retail-presentations-and-pricing.xlsx"
WHOLESALE_XLSX = PRODUCTS_DIR / "forestal-mt-products-wholesale-presentations-and-pricing.xlsx"

CATALOG_JSON = PRODUCTS_DIR / "catalogue.json"
SKU_BASE_JSON = PRODUCTS_DIR / "sku-base.json"
SKU_PRESENTATIONS_JSON = PRODUCTS_DIR / "sku-presentations.json"
RETAIL_CSV = PRODUCTS_DIR / "retail.csv"
WHOLESALE_CSV = PRODUCTS_DIR / "wholesale.csv"

METADATA = OrderedDict(
    repository="forestal-mt-suite",
    owner="nerymurillohn",
    type="Brand and Product Data Repository",
    company="Forestal MT (Forestal Murillo Tejada S. de R.L. de C.V.)",
    country="Honduras",
    industry="Artisan Furniture Manufacturing",
)

CATALOG_HEADERS = (
    "SKU",
    "Product Name",
    "Seal",
    "Collection",
    "Botanical Source",
    "HS Code",
    "Part Used",
    "Product Description",
    "Ingredients",
    "Traditional Uses",
    "Processing Method",
    "Country of Origin",
    "Hero Image Reference",
    "Doc Reference",
)

RETAIL_HEADERS = [
    "SKU",
    "Product Name",
    "Collection",
    "Retail Presentation",
    "Presentation Type",
    "Measurement",
    "Unit",
    "Currency",
    "Price",
]

WHOLESALE_HEADERS = [
    "SKU",
    "Product Name",
    "Collection",
    "Wholesale Presentation",
    "Presentation Type",
    "Measurement",
    "Unit",
    "Currency",
    "Price",
    "Pack Quantity",
    "Pack Unit",
    "Inner Measurement",
    "Inner Unit",
]

CASE_PRESENTATION_RE = re.compile(
    r"^Case \((?P<count>\d+)\s+(?P<pack_unit>[A-Za-z ]+?)\s+x\s+(?P<inner_amount>[0-9]+(?:\.[0-9]+)?)\s+(?P<inner_unit>[A-Za-z]+)\)$",
    re.IGNORECASE,
)

# Some wholesale pack configurations cannot be expressed directly in the Excel
# master (binary diffs are discouraged in the repo), so we maintain the case-pack
# metadata here and overlay it before writing the CSV/JSON artifacts.
WHOLESALE_CASE_PACK_OVERRIDES = {
    "FMT-BO-SB-2025": [
        {
            "wholesale_presentation": "Case (12 bars x 100 g)",
            "presentation_type": "case pack",
            "measurement": 12,
            "unit": "bars",
            "pack_quantity": 12,
            "pack_unit": "bars",
            "inner_measurement": 100,
            "inner_unit": "g",
        },
        {
            "wholesale_presentation": "Case (24 bars x 100 g)",
            "presentation_type": "case pack",
            "measurement": 24,
            "unit": "bars",
            "pack_quantity": 24,
            "pack_unit": "bars",
            "inner_measurement": 100,
            "inner_unit": "g",
        },
        {
            "wholesale_presentation": "Case (48 bars x 100 g)",
            "presentation_type": "case pack",
            "measurement": 48,
            "unit": "bars",
            "pack_quantity": 48,
            "pack_unit": "bars",
            "inner_measurement": 100,
            "inner_unit": "g",
        },
        {
            "wholesale_presentation": "Case (96 bars x 100 g)",
            "presentation_type": "case pack",
            "measurement": 96,
            "unit": "bars",
            "pack_quantity": 96,
            "pack_unit": "bars",
            "inner_measurement": 100,
            "inner_unit": "g",
        },
    ]
}


def _load_rows(path: Path) -> Iterable[Dict[str, object]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(value).strip() if value else "" for value in header_row]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        values = []
        for value in row:
            if isinstance(value, str):
                cleaned = value.replace("\u0007", "•")
                values.append(cleaned.strip())
            else:
                values.append(value)
        data = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        yield data


def build_base_catalog() -> Dict[str, Dict[str, object]]:
    base: Dict[str, Dict[str, object]] = {}
    for row in _load_rows(CATALOG_XLSX):
        sku = str(row["SKU"]).strip()
        if not sku:
            continue
        if sku not in base:
            base[sku] = {
                "sku": sku,
                "product_name": row.get("Product Name", ""),
                "seal": row.get("Seal", ""),
                "collection": row.get("Collection", ""),
                "botanical_source": row.get("Botanical Source", ""),
                "hs_code": int(row["HS Code"]) if isinstance(row.get("HS Code"), (int, float)) else row.get("HS Code"),
                "part_used": row.get("Part Used", ""),
                "product_description": row.get("Product Description", ""),
                "ingredients": row.get("Ingredients", ""),
                "traditional_uses": row.get("Traditional Uses", ""),
                "processing_method": row.get("Processing Method", ""),
                "country_of_origin": row.get("Country of Origin", ""),
                "hero_image_reference": row.get("Hero Image Reference", ""),
                "doc_reference": row.get("Doc Reference", ""),
            }
        else:
            for key in CATALOG_HEADERS[1:]:
                normalized_key = {
                    "Product Name": "product_name",
                    "Seal": "seal",
                    "Collection": "collection",
                    "Botanical Source": "botanical_source",
                    "HS Code": "hs_code",
                    "Part Used": "part_used",
                    "Product Description": "product_description",
                    "Ingredients": "ingredients",
                    "Traditional Uses": "traditional_uses",
                    "Processing Method": "processing_method",
                    "Country of Origin": "country_of_origin",
                    "Hero Image Reference": "hero_image_reference",
                    "Doc Reference": "doc_reference",
                }[key]
                current = base[sku][normalized_key]
                new_value = row.get(key)
                if isinstance(new_value, str):
                    new_value = new_value.replace("\u0007", "•").strip()
                if current in (None, "") and new_value:
                    base[sku][normalized_key] = new_value
                elif new_value and current != new_value:
                    raise ValueError(f"Mismatch for {sku} field {normalized_key}: '{current}' vs '{new_value}'")
    return base


def _normalize_number(value: object) -> Optional[object]:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _apply_wholesale_overrides(rows: List[Dict[str, object]]) -> List[Dict[str, object]]:
    adjusted: List[Dict[str, object]] = []
    remaining_overrides = {sku: overrides.copy() for sku, overrides in WHOLESALE_CASE_PACK_OVERRIDES.items()}
    for row in rows:
        sku = str(row.get("SKU", "")).strip()
        overrides = remaining_overrides.get(sku)
        if overrides:
            override = overrides.pop(0)
            new_row = row.copy()
            presentation = override.get("wholesale_presentation")
            if presentation:
                new_row["Wholesale Presentation"] = presentation
            presentation_type = override.get("presentation_type")
            if presentation_type:
                new_row["Presentation Type"] = presentation_type
            for key in ("Measurement", "Unit", "Currency", "Price"):
                if key in override:
                    new_row[key] = override[key]
            for extra_key, target_key in (
                ("pack_quantity", "Pack Quantity"),
                ("pack_unit", "Pack Unit"),
                ("inner_measurement", "Inner Measurement"),
                ("inner_unit", "Inner Unit"),
            ):
                if extra_key in override:
                    new_row[target_key] = override[extra_key]
            adjusted.append(new_row)
        else:
            adjusted.append(row)
    for sku, overrides in remaining_overrides.items():
        if overrides:
            raise ValueError(
                f"Wholesale overrides for {sku} not fully applied; {len(overrides)} definitions unused"
            )
    return adjusted


def _enrich_wholesale_rows(rows: List[Dict[str, object]]) -> None:
    for row in rows:
        presentation = str(row.get("Wholesale Presentation", "")).strip()
        match = CASE_PRESENTATION_RE.match(presentation)
        if match:
            count_raw = match.group("count")
            pack_unit = match.group("pack_unit").strip()
            inner_amount_raw = match.group("inner_amount")
            inner_unit = match.group("inner_unit").strip()

            if count_raw:
                count = int(float(count_raw))
                row["Pack Quantity"] = count
            if pack_unit:
                row["Pack Unit"] = pack_unit
            if inner_amount_raw:
                inner_amount = float(inner_amount_raw)
                if inner_amount.is_integer():
                    inner_amount = int(inner_amount)
                row["Inner Measurement"] = inner_amount
            if inner_unit:
                row["Inner Unit"] = inner_unit
        else:
            row.setdefault("Pack Quantity", None)
            row.setdefault("Pack Unit", None)
            row.setdefault("Inner Measurement", None)
            row.setdefault("Inner Unit", None)


def _build_presentations(channel: str, headers: List[str], rows: Iterable[Dict[str, object]]) -> List[Dict[str, object]]:
    items: List[Dict[str, object]] = []
    for row in rows:
        sku = str(row.get("SKU", "")).strip()
        if not sku:
            continue
        measurement = _normalize_number(row.get("Measurement"))
        price = _normalize_number(row.get("Price"))
        pack_quantity = _normalize_number(row.get("Pack Quantity"))
        inner_measurement = _normalize_number(row.get("Inner Measurement"))
        item = {
            "sku": sku,
            "channel": channel,
            "presentation": row.get(headers[3], ""),
            "presentation_type": row.get("Presentation Type", ""),
            "measurement": measurement,
            "unit": row.get("Unit", ""),
            "currency": row.get("Currency", ""),
            "price": price,
        }
        if pack_quantity not in (None, ""):
            item["pack_quantity"] = pack_quantity
            pack_unit = row.get("Pack Unit")
            if pack_unit not in (None, ""):
                item["pack_unit"] = pack_unit
        inner_unit = row.get("Inner Unit")
        if inner_measurement not in (None, ""):
            item["inner_measurement"] = inner_measurement
            if inner_unit not in (None, ""):
                item["inner_unit"] = inner_unit
        elif inner_unit not in (None, ""):
            # Preserve explicit unit metadata even if the measurement is blank.
            item["inner_unit"] = inner_unit
        items.append(item)
    items.sort(key=lambda obj: (obj["sku"], obj["channel"], str(obj["presentation"])))
    return items


def write_csv(path: Path, headers: List[str], rows: Iterable[Dict[str, object]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header) for header in headers})


def main() -> None:
    base = build_base_catalog()
    retail_rows = list(_load_rows(RETAIL_XLSX))
    wholesale_rows = list(_load_rows(WHOLESALE_XLSX))
    wholesale_rows = _apply_wholesale_overrides(wholesale_rows)
    _enrich_wholesale_rows(wholesale_rows)
    retail_presentations = _build_presentations("retail", RETAIL_HEADERS, retail_rows)
    wholesale_presentations = _build_presentations("wholesale", WHOLESALE_HEADERS, wholesale_rows)

    write_csv(RETAIL_CSV, RETAIL_HEADERS, retail_rows)
    write_csv(WHOLESALE_CSV, WHOLESALE_HEADERS, wholesale_rows)

    base_list = [base[sku] for sku in sorted(base.keys())]
    presentations = retail_presentations + wholesale_presentations

    for path, payload in (
        (SKU_BASE_JSON, base_list),
        (SKU_PRESENTATIONS_JSON, presentations),
    ):
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    catalogue_payload = OrderedDict(METADATA)
    catalogue_payload["generated_at"] = datetime.now(timezone.utc).isoformat()
    catalogue_payload["products"] = base_list
    catalogue_payload["presentations"] = presentations
    CATALOG_JSON.write_text(json.dumps(catalogue_payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
