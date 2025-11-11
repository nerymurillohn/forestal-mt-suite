#!/usr/bin/env python3
"""Semantic validation tests for Forestal MT Suite product data.

These tests validate data semantics, not just existence:
- HS code format and validity
- Botanical nomenclature (Latin binomial names)
- Required fields per collection
- Unit consistency
- Cross-document parity (Excel ↔ JSON)
"""
import json
import re
from pathlib import Path

import pytest

try:
    import jsonschema
    HAS_JSONSCHEMA = True
except ImportError:
    HAS_JSONSCHEMA = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

ROOT = Path(__file__).resolve().parents[1]


@pytest.mark.skipif(not HAS_JSONSCHEMA, reason="jsonschema not installed")
def test_sku_base_schema_validation():
    """Validate sku-base.json against JSON Schema."""
    schema_path = ROOT / "tests" / "schemas" / "product-schema.json"
    data_path = ROOT / "products-data" / "sku-base.json"

    with open(schema_path) as f:
        schema = json.load(f)

    with open(data_path) as f:
        data = json.load(f)

    # This will raise ValidationError if data doesn't match schema
    jsonschema.validate(instance=data, schema=schema)


def test_hs_code_format():
    """Validate HS codes are 6-10 digits and properly formatted."""
    data_path = ROOT / "products-data" / "sku-base.json"

    with open(data_path) as f:
        products = json.load(f)

    invalid_hs_codes = []
    for product in products:
        hs_code = str(product.get("hs_code", ""))

        # HS codes must be 6-10 digits
        if not re.match(r'^\d{6,10}$', hs_code):
            invalid_hs_codes.append({
                "sku": product.get("sku"),
                "hs_code": hs_code,
                "reason": "Must be 6-10 digits"
            })

    assert len(invalid_hs_codes) == 0, f"Invalid HS codes found: {invalid_hs_codes}"


def test_botanical_nomenclature():
    """Validate botanical names follow Latin binomial nomenclature."""
    data_path = ROOT / "products-data" / "sku-base.json"

    with open(data_path) as f:
        products = json.load(f)

    # Pattern: Genus species (e.g., "Elaeis oleifera")
    #          Genus spp. (multiple species, e.g., "Amaranthus spp.")
    #          "Multiple sources" (for herb blends)
    pattern = re.compile(r'^[A-Z][a-z]+\s([a-z]+|spp\.)$|^Multiple sources$')

    invalid_names = []
    for product in products:
        botanical_source = product.get("botanical_source", "")

        if not pattern.match(botanical_source):
            invalid_names.append({
                "sku": product.get("sku"),
                "botanical_source": botanical_source,
                "reason": "Must follow 'Genus species', 'Genus spp.', or 'Multiple sources'"
            })

    assert len(invalid_names) == 0, f"Invalid botanical names: {invalid_names}"


def test_country_of_origin_consistency():
    """All products must have 'Honduras' as country of origin."""
    data_path = ROOT / "products-data" / "sku-base.json"

    with open(data_path) as f:
        products = json.load(f)

    non_honduras = []
    for product in products:
        country = product.get("country_of_origin", "")

        if country != "Honduras":
            non_honduras.append({
                "sku": product.get("sku"),
                "country_of_origin": country
            })

    assert len(non_honduras) == 0, f"Products with non-Honduras origin: {non_honduras}"


def test_required_fields_per_product():
    """Verify all required fields are present and non-empty."""
    data_path = ROOT / "products-data" / "sku-base.json"

    with open(data_path) as f:
        products = json.load(f)

    required_fields = [
        "sku",
        "product_name",
        "collection",
        "botanical_source",
        "hs_code",
        "ingredients",
        "country_of_origin",
        "hero_image_reference",
        "doc_reference"
    ]

    missing_fields = []
    for product in products:
        sku = product.get("sku", "UNKNOWN")

        for field in required_fields:
            if field not in product:
                missing_fields.append({
                    "sku": sku,
                    "missing_field": field
                })
            elif not product[field]:  # Check for empty values
                missing_fields.append({
                    "sku": sku,
                    "empty_field": field
                })

    assert len(missing_fields) == 0, f"Missing/empty required fields: {missing_fields}"


def test_collection_specific_requirements():
    """Validate collection-specific field requirements."""
    data_path = ROOT / "products-data" / "sku-base.json"

    with open(data_path) as f:
        products = json.load(f)

    issues = []

    for product in products:
        sku = product.get("sku")
        collection = product.get("collection")

        # Batana Oil products should have "Elaeis oleifera" as botanical source
        if collection == "Batana Oil":
            if product.get("botanical_source") != "Elaeis oleifera":
                issues.append({
                    "sku": sku,
                    "issue": "Batana Oil products must use 'Elaeis oleifera'",
                    "actual": product.get("botanical_source")
                })

        # Stingless Bee Honey should have "Tetragonisca angustula" or similar
        elif collection == "Stingless Bee Honey":
            botanical = product.get("botanical_source", "")
            if not ("Tetragonisca" in botanical or "Melipona" in botanical):
                issues.append({
                    "sku": sku,
                    "issue": "Stingless Bee Honey should reference Tetragonisca or Melipona genus",
                    "actual": botanical
                })

        # Traditional Herbs can have "Multiple sources"
        elif collection == "Traditional Herbs":
            # No strict requirement, but should have valid botanical name or "Multiple sources"
            pass

    assert len(issues) == 0, f"Collection-specific validation failed: {issues}"


def test_sku_naming_consistency():
    """Validate SKU names match expected patterns for each collection."""
    data_path = ROOT / "products-data" / "sku-base.json"

    with open(data_path) as f:
        products = json.load(f)

    collection_prefixes = {
        "Batana Oil": "FMT-BO-",
        "Stingless Bee Honey": "FMT-SBH-",
        "Traditional Herbs": "FMT-TH-"
    }

    mismatched = []
    for product in products:
        sku = product.get("sku", "")
        collection = product.get("collection", "")

        expected_prefix = collection_prefixes.get(collection)
        if expected_prefix and not sku.startswith(expected_prefix):
            mismatched.append({
                "sku": sku,
                "collection": collection,
                "expected_prefix": expected_prefix
            })

    assert len(mismatched) == 0, f"SKU prefixes don't match collections: {mismatched}"


def test_hero_image_path_consistency():
    """Validate hero image paths match SKU naming convention."""
    data_path = ROOT / "products-data" / "sku-base.json"

    with open(data_path) as f:
        products = json.load(f)

    mismatched = []
    for product in products:
        sku = product.get("sku", "")
        hero_path = product.get("hero_image_reference", "")

        # Expected: assets/products-images/{collection-slug}/hero/{sku-lowercase}-hero.png
        expected_filename = f"{sku.lower()}-hero.png"

        if expected_filename not in hero_path.lower():
            mismatched.append({
                "sku": sku,
                "hero_path": hero_path,
                "expected_filename": expected_filename
            })

    assert len(mismatched) == 0, f"Hero image paths don't match SKUs: {mismatched}"


def test_ingredient_list_format():
    """Validate ingredient lists are properly formatted."""
    data_path = ROOT / "products-data" / "sku-base.json"

    with open(data_path) as f:
        products = json.load(f)

    issues = []
    for product in products:
        sku = product.get("sku")
        ingredients = product.get("ingredients", "")

        # Ingredients should be at least 10 characters (meaningful content)
        if len(ingredients) < 10:
            issues.append({
                "sku": sku,
                "issue": "Ingredients too short (< 10 chars)",
                "length": len(ingredients)
            })

        # Should not contain obvious placeholder text
        placeholders = ["lorem", "ipsum", "placeholder", "TODO", "TBD"]
        if any(placeholder in ingredients.lower() for placeholder in placeholders):
            issues.append({
                "sku": sku,
                "issue": "Ingredients contain placeholder text",
                "ingredients": ingredients[:50]
            })

    assert len(issues) == 0, f"Ingredient format issues: {issues}"


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
def test_excel_json_parity():
    """Validate Excel master data matches JSON output (sku-base.json)."""
    excel_path = ROOT / "products-data" / "forestal-mt-products-catalogue-46-skus.xlsx"
    json_path = ROOT / "products-data" / "sku-base.json"

    # Load Excel
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    # Find header row (assuming row 1)
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)

    # Extract Excel SKUs
    excel_skus = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Assuming SKU is first column
            excel_skus.add(str(row[0]).strip())

    # Load JSON
    with open(json_path) as f:
        json_products = json.load(f)

    json_skus = {p["sku"] for p in json_products}

    # Check parity
    missing_in_json = excel_skus - json_skus
    missing_in_excel = json_skus - excel_skus

    issues = []
    if missing_in_json:
        issues.append(f"SKUs in Excel but not in JSON: {missing_in_json}")
    if missing_in_excel:
        issues.append(f"SKUs in JSON but not in Excel: {missing_in_excel}")

    assert len(issues) == 0, f"Excel ↔ JSON parity issues: {issues}"


def test_no_duplicate_skus():
    """Ensure no duplicate SKUs exist in sku-base.json."""
    data_path = ROOT / "products-data" / "sku-base.json"

    with open(data_path) as f:
        products = json.load(f)

    skus = [p.get("sku") for p in products]
    duplicates = [sku for sku in set(skus) if skus.count(sku) > 1]

    assert len(duplicates) == 0, f"Duplicate SKUs found: {duplicates}"


def test_no_orphaned_collections():
    """Verify all collections referenced in products have documentation."""
    data_path = ROOT / "products-data" / "sku-base.json"
    docs_path = ROOT / "docs" / "collections"

    with open(data_path) as f:
        products = json.load(f)

    # Get unique collections from products
    collections = {p.get("collection") for p in products}

    # Map collections to expected doc files
    collection_to_file = {
        "Batana Oil": "batana-oil.md",
        "Stingless Bee Honey": "stingless-bee-honey.md",
        "Traditional Herbs": "traditional-herbs.md"
    }

    missing_docs = []
    for collection in collections:
        expected_file = collection_to_file.get(collection)
        if expected_file:
            doc_path = docs_path / expected_file
            if not doc_path.exists():
                missing_docs.append({
                    "collection": collection,
                    "expected_file": expected_file
                })

    assert len(missing_docs) == 0, f"Missing collection documentation: {missing_docs}"
